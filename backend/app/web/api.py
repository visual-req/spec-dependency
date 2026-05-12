from __future__ import annotations

import json
import io
import sys
import time
import urllib.request
from pathlib import Path

from starlette.requests import Request

from ..config.config import load_config
from ..io.docx_text_extractor import extract_word_text
from ..io.excel_exporter import write_module_dependency_excel
from ..llm.llm_client import llm_adjust_scan_result, llm_extract_dependency_catalog_from_excel, llm_extract_module_deps
from ..core.dependency_catalog import DependencyCatalog
from ..core.module_splitter import split_text_into_modules


def create_app():
    try:
        from fastapi import FastAPI, File, UploadFile
        from fastapi.responses import FileResponse, JSONResponse
        from fastapi.staticfiles import StaticFiles
    except Exception as e:
        raise RuntimeError("fastapi is required to run the web server") from e

    globals()["UploadFile"] = UploadFile
    globals()["File"] = File

    app = FastAPI()
    cfg = load_config()
    deps_json = _resolve_deps_json(cfg)
    catalog = DependencyCatalog.load(deps_json)

    project_root = Path(__file__).resolve().parents[3]
    bundle_root = Path(getattr(sys, "_MEIPASS")).resolve() if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS") else None
    dist = (bundle_root / "frontend" / "dist") if bundle_root is not None else (project_root / "frontend" / "dist")
    assets = dist / "assets"
    if assets.is_dir():
        app.mount("/assets", StaticFiles(directory=str(assets)), name="assets")

    def append_log(kind: str, payload: dict) -> None:
        ts = time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime())
        base = cfg.work_dir if cfg.work_dir is not None else (Path.cwd() / "work").resolve()
        log_dir = (base / "logs").resolve()
        log_dir.mkdir(parents=True, exist_ok=True)
        path = log_dir / "spec_dep.log.jsonl"
        record = {"ts": ts, "kind": kind, "data": payload}
        try:
            with path.open("a", encoding="utf-8") as f:
                f.write(json.dumps(record, ensure_ascii=False) + "\n")
        except Exception:
            return

    def _dedupe_modules(modules) -> list:
        ms = modules if isinstance(modules, list) else []
        out = []
        for m in ms:
            if not isinstance(m, dict):
                continue
            name = m.get("module") or "global"
            deps = m.get("dependencies")
            deps_list = deps if isinstance(deps, list) else []
            sys_map: dict[str, dict] = {}
            sys_order: list[str] = []
            for d in deps_list:
                if not isinstance(d, dict):
                    continue
                sid = d.get("id")
                sname = d.get("name")
                key = (str(sid).strip() if sid is not None else "") or (str(sname).strip() if sname is not None else "")
                if not key:
                    continue
                if key not in sys_map:
                    sys_map[key] = dict(d)
                    sys_map[key]["apis"] = []
                    sys_order.append(key)
                cur = sys_map[key]
                cur["confidence"] = _better_conf(cur.get("confidence"), d.get("confidence"))
                for k in ["id", "name", "type", "reason", "evidence"]:
                    if not cur.get(k) and d.get(k):
                        cur[k] = d.get(k)
                apis = d.get("apis")
                api_list = apis if isinstance(apis, list) else []
                api_map = { _api_key(a): a for a in (cur.get("apis") or []) if isinstance(a, dict) and _api_key(a) }
                for a in api_list:
                    if not isinstance(a, dict):
                        continue
                    ak = _api_key(a)
                    if not ak:
                        continue
                    if ak not in api_map:
                        api_map[ak] = a
                    else:
                        api_map[ak] = _better_api(api_map[ak], a)
                cur["apis"] = list(api_map.values())
            out.append({"module": name, "dependencies": [sys_map[k] for k in sys_order]})
        return out

    def _api_key(a: dict) -> str:
        aid = a.get("id")
        name = a.get("name")
        method = a.get("method")
        path = a.get("path")
        first = (str(aid).strip() if aid is not None else "") or (str(name).strip() if name is not None else "")
        return "|".join([first, str(method or ""), str(path or "")]).strip("|")

    def _better_conf(a, b):
        try:
            na = float(a)
        except Exception:
            na = None
        try:
            nb = float(b)
        except Exception:
            nb = None
        if nb is not None and (na is None or nb > na):
            return b
        return a

    def _better_api(old, new) -> dict:
        if not isinstance(old, dict):
            return dict(new)
        if not isinstance(new, dict):
            return dict(old)
        out = dict(old)
        out["confidence"] = _better_conf(out.get("confidence"), new.get("confidence"))
        for k in ["id", "name", "method", "path", "reason", "evidence"]:
            if not out.get(k) and new.get(k):
                out[k] = new.get(k)
        return out

    def _resolve_scan_dir(scan_id: str) -> Path | None:
        sid = (scan_id or "").strip()
        if not sid:
            return None
        candidates: list[Path] = []
        if cfg.work_dir is not None:
            candidates.append(cfg.work_dir.resolve())
        candidates.append((project_root / "work").resolve())
        candidates.append((project_root / "executable" / "work").resolve())

        seen: set[str] = set()
        for wd in candidates:
            key = str(wd)
            if key in seen:
                continue
            seen.add(key)
            out_dir = (wd / "output" / sid).resolve()
            scan_json = (out_dir / "scan.json").resolve()
            if scan_json.is_file():
                return out_dir
        return None

    def _as_str(v) -> str:
        if v is None:
            return ""
        if isinstance(v, str):
            return v
        return str(v)

    def _snake_upper(s: str) -> str:
        t = (s or "").strip()
        if not t:
            return ""
        out = []
        prev_us = False
        for ch in t:
            is_alnum = ("0" <= ch <= "9") or ("a" <= ch <= "z") or ("A" <= ch <= "Z")
            if is_alnum:
                out.append(ch.upper())
                prev_us = False
            else:
                if not prev_us:
                    out.append("_")
                    prev_us = True
        r = "".join(out).strip("_")
        while "__" in r:
            r = r.replace("__", "_")
        if not r:
            return ""
        if not (("A" <= r[0] <= "Z") or r[0] == "_"):
            r = "_" + r
        return r

    def _keywords_from(parts: list[str], limit: int = 10) -> list[str]:
        seen: set[str] = set()
        out: list[str] = []
        for p in parts:
            for raw in (p or "").replace("/", " ").replace("-", " ").replace("_", " ").split():
                w = raw.strip()
                if not w:
                    continue
                if w in seen:
                    continue
                seen.add(w)
                out.append(w)
                if len(out) >= limit:
                    return out
        return out

    def _fetch_openapi(url: str, extra_headers: dict | None = None) -> dict | None:
        u = (url or "").strip()
        if not (u.startswith("http://") or u.startswith("https://")):
            return None
        headers = {"User-Agent": "spec-dep/1.0"}
        if isinstance(extra_headers, dict):
            for k, v in extra_headers.items():
                kk = _as_str(k).strip()
                vv = _as_str(v).strip()
                if kk and vv:
                    headers[kk] = vv
        req = urllib.request.Request(u, headers=headers, method="GET")
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read(6 * 1024 * 1024).decode("utf-8", errors="replace")
        except Exception:
            return None
        try:
            obj = json.loads(raw)
            return obj if isinstance(obj, dict) else None
        except Exception:
            pass
        try:
            import yaml
        except Exception:
            return None
        try:
            loaded = yaml.safe_load(raw)
        except Exception:
            return None
        return loaded if isinstance(loaded, dict) else None

    def _build_fetch_headers(body) -> dict:
        token = _as_str(body.get("token")).strip()
        if not token:
            return {}
        header_name = _as_str(body.get("token_header") or "Authorization").strip() or "Authorization"
        scheme = _as_str(body.get("token_scheme") or "bearer").strip().lower()
        if scheme == "bearer":
            value = token if token.lower().startswith("bearer ") else ("Bearer " + token)
        elif scheme == "raw":
            value = token
        else:
            value = token
        return {header_name: value}

    def _extract_openapi_apis(spec: dict) -> list[dict]:
        paths = spec.get("paths")
        if not isinstance(paths, dict):
            return []
        methods = {"get", "post", "put", "delete", "patch", "head", "options"}
        out: list[dict] = []
        for p, v in paths.items():
            if not isinstance(v, dict):
                continue
            path = _as_str(p).strip()
            for m, op in v.items():
                mm = _as_str(m).lower().strip()
                if mm not in methods:
                    continue
                if not isinstance(op, dict):
                    continue
                summary = _as_str(op.get("summary")).strip()
                op_id = _as_str(op.get("operationId")).strip()
                tags = op.get("tags")
                tag_list = [str(x).strip() for x in tags] if isinstance(tags, list) else []
                name = summary or op_id or (mm.upper() + " " + path)
                api_id = _snake_upper(op_id) if op_id else _snake_upper(mm.upper() + "_" + path)
                segs = [x for x in path.split("/") if x]
                keywords = _keywords_from([name] + tag_list + segs + [path])
                out.append({"id": api_id or _snake_upper(name), "name": name, "method": mm.upper(), "path": path, "keywords": keywords})
        out = [a for a in out if isinstance(a.get("id"), str) and a["id"].strip() and isinstance(a.get("name"), str) and a["name"].strip()]
        seen: set[str] = set()
        deduped: list[dict] = []
        for a in out:
            k = a.get("id") or ""
            if k in seen:
                continue
            seen.add(k)
            deduped.append(a)
        return deduped

    def _extract_yapi_apis(obj: dict) -> list[dict]:
        root = obj if isinstance(obj, dict) else {}
        cats = None
        if isinstance(root.get("cats"), list):
            cats = root.get("cats")
        elif isinstance(root.get("data"), list) and ("errcode" in root or "errmsg" in root):
            cats = root.get("data")
        if cats is None and isinstance(root.get("interfaces"), list):
            cats = [{"name": root.get("name") or "default", "list": root.get("interfaces")}]
        if not isinstance(cats, list):
            return []

        out: list[dict] = []
        for c in cats:
            if not isinstance(c, dict):
                continue
            cname = _as_str(c.get("name") or c.get("title") or "").strip()
            lst = c.get("list")
            if not isinstance(lst, list):
                continue
            for it in lst:
                if not isinstance(it, dict):
                    continue
                path = _as_str(it.get("path") or it.get("url") or "").strip()
                method = _as_str(it.get("method") or "").strip().upper()
                title = _as_str(it.get("title") or it.get("name") or "").strip()
                iid = it.get("_id")
                if iid is None:
                    iid = it.get("id")
                api_id = ""
                if iid is not None and str(iid).strip():
                    api_id = _snake_upper(f"YAPI_{iid}")
                if not api_id:
                    api_id = _snake_upper((method or "GET") + "_" + path)
                name = title or (method + " " + path if method and path else api_id)
                if not path:
                    continue
                if method not in {"GET", "POST", "PUT", "DELETE", "PATCH", "HEAD", "OPTIONS"}:
                    method = "GET"
                segs = [x for x in path.split("/") if x]
                keywords = _keywords_from([name, cname] + segs + [path])
                out.append({"id": api_id or _snake_upper(name), "name": name, "method": method, "path": path, "keywords": keywords})

        out = [a for a in out if isinstance(a.get("id"), str) and a["id"].strip() and isinstance(a.get("name"), str) and a["name"].strip()]
        seen: set[str] = set()
        deduped: list[dict] = []
        for a in out:
            k = a.get("id") or ""
            if k in seen:
                continue
            seen.add(k)
            deduped.append(a)
        return deduped

    def _merge_apis(existing: list[dict], incoming: list[dict]) -> tuple[list[dict], int, int]:
        ex = [x for x in existing if isinstance(x, dict)]
        inc = [x for x in incoming if isinstance(x, dict)]
        m: dict[str, dict] = {}
        order: list[str] = []
        for a in ex:
            aid = _as_str(a.get("id")).strip()
            if not aid:
                continue
            if aid not in m:
                m[aid] = dict(a)
                order.append(aid)
        added = 0
        updated = 0
        for a in inc:
            aid = _as_str(a.get("id")).strip()
            if not aid:
                continue
            if aid not in m:
                m[aid] = dict(a)
                order.append(aid)
                added += 1
                continue
            cur = m[aid]
            for k in ["name", "method", "path"]:
                nv = a.get(k)
                if nv is not None and str(nv).strip() and str(cur.get(k) or "").strip() != str(nv).strip():
                    cur[k] = nv
                    updated += 1
            if a.get("deleted") is True and cur.get("deleted") is not True:
                cur["deleted"] = True
                updated += 1
            if a.get("deleted") is False and cur.get("deleted") is True:
                cur["deleted"] = False
                updated += 1
            for k in ["request_params", "response_params"]:
                nv = a.get(k)
                if isinstance(nv, list) and nv != (cur.get(k) if isinstance(cur.get(k), list) else None):
                    cur[k] = nv
                    updated += 1
            kw_old = cur.get("keywords") if isinstance(cur.get("keywords"), list) else []
            kw_new = a.get("keywords") if isinstance(a.get("keywords"), list) else []
            merged_kw = []
            seen_kw: set[str] = set()
            for x in [*kw_old, *kw_new]:
                s = _as_str(x).strip()
                if not s or s in seen_kw:
                    continue
                seen_kw.add(s)
                merged_kw.append(s)
            cur["keywords"] = merged_kw
        return ([m[k] for k in order], added, updated)

    def _merge_list_str(old, new) -> list[str]:
        a = old if isinstance(old, list) else []
        b = new if isinstance(new, list) else []
        seen: set[str] = set()
        out: list[str] = []
        for x in [*a, *b]:
            s = _as_str(x).strip()
            if not s or s in seen:
                continue
            seen.add(s)
            out.append(s)
        return out

    def _normalize_method(v) -> str | None:
        m = _as_str(v).strip().upper()
        if not m:
            return None
        if m not in {"GET", "POST", "PUT", "DELETE", "PATCH", "HEAD", "OPTIONS"}:
            return None
        return m

    def _normalize_params_list(v) -> list[dict]:
        if not isinstance(v, list):
            return []
        out: list[dict] = []
        for it in v:
            if not isinstance(it, dict):
                continue
            name = _as_str(it.get("name")).strip()
            if not name:
                continue
            item = {"name": name}
            typ = _as_str(it.get("type")).strip()
            if typ:
                item["type"] = typ
            desc = _as_str(it.get("desc") or it.get("description")).strip()
            if desc:
                item["desc"] = desc
            if it.get("required") is True:
                item["required"] = True
            out.append(item)
        return out

    def _normalize_api_item(a: dict) -> dict | None:
        if not isinstance(a, dict):
            return None
        name = _as_str(a.get("name")).strip()
        method = _normalize_method(a.get("method"))
        path = _as_str(a.get("path")).strip()
        if not name and not path:
            return None
        aid = _as_str(a.get("id")).strip()
        if not aid:
            seed = name or ((method or "API") + "_" + path) if path else name
            aid = _snake_upper(seed)
        if not aid:
            return None
        out = {"id": aid, "name": name or aid}
        if method is not None:
            out["method"] = method
        if path:
            out["path"] = path
        kws = a.get("keywords") if isinstance(a.get("keywords"), list) else []
        if kws:
            out["keywords"] = _merge_list_str([], kws)
        if a.get("deleted") is True:
            out["deleted"] = True
        rp = _normalize_params_list(a.get("request_params"))
        if rp:
            out["request_params"] = rp
        sp = _normalize_params_list(a.get("response_params"))
        if sp:
            out["response_params"] = sp
        return out

    def _load_deps_root() -> dict:
        try:
            root = json.loads(deps_json.read_text(encoding="utf-8"))
        except Exception:
            root = {}
        if not isinstance(root, dict):
            root = {}
        systems = root.get("systems")
        if not isinstance(systems, list):
            root["systems"] = []
        return root

    def _save_deps_root(root: dict) -> None:
        deps_json.write_text(json.dumps(root, ensure_ascii=False, indent=2), encoding="utf-8")

    def _read_api_file_obj(api_file_name: str) -> dict:
        dep_dir = deps_json.parent.resolve()
        p = (dep_dir / api_file_name).resolve()
        if not p.is_file():
            return {}
        try:
            obj = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            obj = {}
        return obj if isinstance(obj, dict) else {}

    def _write_api_file_obj(api_file_name: str, sys_id: str, apis: list[dict]) -> str:
        dep_dir = deps_json.parent.resolve()
        p = (dep_dir / api_file_name).resolve()
        p.write_text(json.dumps({"system_id": sys_id, "apis": apis}, ensure_ascii=False, indent=2), encoding="utf-8")
        return str(p)

    def _known_systems_for_llm(limit: int = 200) -> list[dict]:
        root = _load_deps_root()
        systems = root.get("systems") if isinstance(root.get("systems"), list) else []
        out: list[dict] = []
        for s in systems:
            if not isinstance(s, dict):
                continue
            sid = _as_str(s.get("id")).strip()
            name = _as_str(s.get("name")).strip()
            if not sid or not name:
                continue
            out.append(
                {
                    "id": sid,
                    "name": name,
                    "type": _as_str(s.get("type")).strip() or "external_service",
                    "aliases": s.get("aliases") if isinstance(s.get("aliases"), list) else [],
                    "keywords": s.get("keywords") if isinstance(s.get("keywords"), list) else [],
                    "description": _as_str(s.get("description")).strip(),
                    "deleted": s.get("deleted") is True,
                }
            )
            if len(out) >= limit:
                break
        return out

    def _upsert_system_root(
        root: dict,
        sys_id: str,
        sys_name: str,
        sys_type: str,
        api_file_name: str,
        extra: dict | None = None,
        replace_lists: bool = False,
    ) -> dict:
        systems = root.get("systems")
        if not isinstance(systems, list):
            systems = []
        found = None
        for s in systems:
            if isinstance(s, dict) and _as_str(s.get("id")).strip() == sys_id:
                found = s
                break
        if found is None:
            found = {"id": sys_id, "name": sys_name, "type": sys_type}
            systems.append(found)
        if sys_name and _as_str(found.get("name")).strip() != sys_name:
            found["name"] = sys_name
        if sys_type and _as_str(found.get("type")).strip() != sys_type:
            found["type"] = sys_type
        found["api_file"] = api_file_name
        if isinstance(extra, dict):
            if extra.get("deleted") is True:
                found["deleted"] = True
            elif "deleted" in extra and extra.get("deleted") is False:
                found["deleted"] = False
            if "description" in extra:
                desc = _as_str(extra.get("description")).strip()
                if desc:
                    found["description"] = desc
                else:
                    if "description" in found:
                        found.pop("description", None)
            aliases = extra.get("aliases") if isinstance(extra.get("aliases"), list) else None
            if aliases is not None:
                if replace_lists:
                    found["aliases"] = _merge_list_str([], aliases)
                else:
                    found["aliases"] = _merge_list_str(found.get("aliases"), aliases)
            keywords = extra.get("keywords") if isinstance(extra.get("keywords"), list) else None
            if keywords is not None:
                if replace_lists:
                    found["keywords"] = _merge_list_str([], keywords)
                else:
                    found["keywords"] = _merge_list_str(found.get("keywords"), keywords)
        root["systems"] = systems
        return found

    def _log_done_urls_for_resume(resume_key: str) -> set[str]:
        key = (resume_key or "").strip()
        if not key:
            return set()
        base = cfg.work_dir if cfg.work_dir is not None else (Path.cwd() / "work").resolve()
        log_path = (base / "logs" / "spec_dep.log.jsonl").resolve()
        if not log_path.is_file():
            return set()
        done: set[str] = set()
        try:
            with log_path.open("r", encoding="utf-8") as f:
                for line in f:
                    s = (line or "").strip()
                    if not s:
                        continue
                    try:
                        rec = json.loads(s)
                    except Exception:
                        continue
                    if not isinstance(rec, dict) or rec.get("kind") != "import_swagger_batch_item":
                        continue
                    data = rec.get("data")
                    if not isinstance(data, dict):
                        continue
                    if _as_str(data.get("resume_key")).strip() != key:
                        continue
                    if data.get("ok") is True:
                        u = _as_str(data.get("url")).strip()
                        if u:
                            done.add(u)
        except Exception:
            return set()
        return done

    @app.get("/api/health")
    def health():
        return {"status": "ok"}

    @app.get("/api/config")
    def config_info():
        return {
            "config_path": str(cfg.config_path) if cfg.config_path is not None else None,
            "work_dir": str(cfg.work_dir) if cfg.work_dir is not None else None,
            "server_host": cfg.server_host,
            "server_port": cfg.server_port,
            "deepseek_base_url": cfg.deepseek_base_url,
            "deepseek_model": cfg.deepseek_model,
            "deepseek_api_key_present": bool(cfg.deepseek_api_key and cfg.deepseek_api_key.strip()),
        }

    @app.get("/api/dependencies")
    def dependencies():
        try:
            raw = deps_json.read_text(encoding="utf-8")
            obj = json.loads(raw) if raw.strip() else {}
        except Exception:
            obj = {}
        desc_map: dict[str, str] = {}
        if isinstance(obj, dict) and isinstance(obj.get("systems"), list):
            for s in obj.get("systems"):
                if not isinstance(s, dict):
                    continue
                sid = s.get("id")
                if not isinstance(sid, str) or not sid.strip():
                    continue
                desc = s.get("description")
                if isinstance(desc, str) and desc.strip():
                    desc_map[sid.strip()] = desc.strip()

        systems_out = []
        for sys in catalog.systems:
            item = {
                "id": sys.id,
                "name": sys.name,
                "type": sys.type,
                "aliases": sys.aliases,
                "keywords": sys.keywords,
                "apis": [{"id": a.id, "name": a.name, "method": a.method, "path": a.path, "keywords": a.keywords} for a in sys.apis],
            }
            if sys.id in desc_map:
                item["description"] = desc_map[sys.id]
            systems_out.append(item)
        return {"systems": systems_out, "file": str(deps_json)}

    @app.get("/api/dependencies_raw")
    def dependencies_raw():
        root = _load_deps_root()
        dep_dir = deps_json.parent.resolve()
        systems = root.get("systems") if isinstance(root.get("systems"), list) else []
        out: list[dict] = []
        for s in systems:
            if not isinstance(s, dict):
                continue
            item = dict(s)
            apis: list[dict] = []
            if isinstance(s.get("apis"), list):
                apis.extend([a for a in s.get("apis") if isinstance(a, dict)])
            api_file = _as_str(s.get("api_file") or s.get("apis_file")).strip()
            if api_file:
                p = (dep_dir / api_file).resolve()
                if p.is_file():
                    try:
                        obj = json.loads(p.read_text(encoding="utf-8"))
                    except Exception:
                        obj = {}
                    if isinstance(obj, dict) and isinstance(obj.get("apis"), list):
                        apis.extend([a for a in obj.get("apis") if isinstance(a, dict)])
                    elif isinstance(obj, list):
                        apis.extend([a for a in obj if isinstance(a, dict)])
            item["apis"] = apis
            out.append(item)
        return {"systems": out, "file": str(deps_json)}

    @app.post("/api/save_dependency_system")
    async def save_dependency_system(request: Request):
        nonlocal catalog
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}

        sys = body.get("system")
        if not isinstance(sys, dict):
            return JSONResponse({"error": "system is required"}, status_code=400)

        sys_name = _as_str(sys.get("name")).strip()
        sys_id = _snake_upper(_as_str(sys.get("id")).strip() or sys_name)
        if not sys_id:
            return JSONResponse({"error": "failed to derive system_id"}, status_code=400)
        if not sys_name:
            sys_name = sys_id
        sys_type = _as_str(sys.get("type")).strip() or "external_service"
        api_file_name = _as_str(sys.get("api_file")).strip() or f"{sys_id}.apis.json"

        apis_in = body.get("apis")
        apis_list = apis_in if isinstance(apis_in, list) else []
        normalized_apis: list[dict] = []
        for a in apis_list:
            na = _normalize_api_item(a) if isinstance(a, dict) else None
            if na is not None:
                normalized_apis.append(na)

        existing = _read_api_file_obj(api_file_name)
        existing_apis = existing.get("apis") if isinstance(existing.get("apis"), list) else []
        merged_apis, added, updated = _merge_apis(existing_apis, normalized_apis)
        api_path = _write_api_file_obj(api_file_name, sys_id, merged_apis)

        root = _load_deps_root()
        extra = {
            "description": sys.get("description"),
            "aliases": sys.get("aliases") if isinstance(sys.get("aliases"), list) else [],
            "keywords": sys.get("keywords") if isinstance(sys.get("keywords"), list) else [],
            "deleted": sys.get("deleted") is True,
        }
        found = _upsert_system_root(root, sys_id, sys_name, sys_type, api_file_name, extra, replace_lists=True)
        _save_deps_root(root)

        catalog = DependencyCatalog.load(deps_json)
        append_log("save_dependency_system", {"system_id": sys_id, "api_count": len(merged_apis), "added": added, "updated": updated, "deleted": sys.get("deleted") is True})
        return {"ok": True, "system": {"id": sys_id, "name": found.get("name"), "type": found.get("type"), "api_file": api_file_name, "deleted": found.get("deleted") is True}, "api_count": len(merged_apis), "added": added, "updated": updated, "deps_file": str(deps_json), "api_file": api_path}

    @app.post("/api/import_deps_excel")
    async def import_deps_excel(request: Request, file: UploadFile = File(...)):
        nonlocal catalog
        if cfg.deepseek_api_key is None or not cfg.deepseek_api_key.strip():
            return JSONResponse({"error": "AI 不可用（请检查 DEEPSEEK_API_KEY 或模型服务）"}, status_code=400)
        try:
            raw_bytes = await file.read()
        except Exception:
            raw_bytes = b""
        if not raw_bytes:
            return JSONResponse({"error": "empty excel file"}, status_code=400)
        try:
            from openpyxl import load_workbook
        except Exception:
            return JSONResponse({"error": "openpyxl is required to read excel files"}, status_code=500)

        try:
            wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
        except Exception:
            return JSONResponse({"error": "failed to parse excel file"}, status_code=400)

        sheets_out: list[dict] = []
        max_rows_per_sheet = 200
        max_cols_per_row = 30
        for ws in wb.worksheets[:10]:
            rows: list[list[str]] = []
            truncated = False
            rcount = 0
            for row in ws.iter_rows(values_only=True):
                rcount += 1
                if rcount > max_rows_per_sheet:
                    truncated = True
                    break
                vals = []
                if row is None:
                    continue
                for c in row[:max_cols_per_row]:
                    s = _as_str(c).strip()
                    vals.append(s)
                while vals and not vals[-1]:
                    vals.pop()
                if not any(vals):
                    continue
                rows.append(vals)
            if rows:
                sheets_out.append({"name": ws.title, "rows": rows, "truncated": truncated})

        excel_json = {"file_name": _as_str(file.filename).strip(), "sheets": sheets_out}
        if not sheets_out:
            return JSONResponse({"error": "no data found in excel"}, status_code=400)

        known = _known_systems_for_llm()
        parsed = llm_extract_dependency_catalog_from_excel(cfg, excel_json, known)
        if parsed is None:
            return JSONResponse({"error": "failed to parse excel via llm"}, status_code=400)

        systems_in = parsed.get("systems")
        systems_list = systems_in if isinstance(systems_in, list) else []
        root = _load_deps_root()
        results: list[dict] = []
        for s in systems_list:
            if not isinstance(s, dict):
                continue
            sys_name = _as_str(s.get("name")).strip()
            sys_id = _snake_upper(_as_str(s.get("id")).strip() or sys_name)
            if not sys_id:
                continue
            if not sys_name:
                sys_name = sys_id
            sys_type = _as_str(s.get("type")).strip() or "external_service"
            api_file_name = f"{sys_id}.apis.json"

            apis_in = s.get("apis")
            apis_list = apis_in if isinstance(apis_in, list) else []
            normalized_apis: list[dict] = []
            for a in apis_list:
                na = _normalize_api_item(a) if isinstance(a, dict) else None
                if na is not None:
                    normalized_apis.append(na)

            existing = _read_api_file_obj(api_file_name)
            existing_apis = existing.get("apis") if isinstance(existing.get("apis"), list) else []
            merged_apis, added, updated = _merge_apis(existing_apis, normalized_apis)
            api_path = _write_api_file_obj(api_file_name, sys_id, merged_apis)

            extra = {
                "description": s.get("description"),
                "aliases": s.get("aliases") if isinstance(s.get("aliases"), list) else [],
                "keywords": s.get("keywords") if isinstance(s.get("keywords"), list) else [],
                "deleted": s.get("deleted") is True,
            }
            found = _upsert_system_root(root, sys_id, sys_name, sys_type, api_file_name, extra)
            results.append(
                {
                    "id": sys_id,
                    "name": found.get("name"),
                    "type": found.get("type"),
                    "api_file": api_file_name,
                    "api_count": len(merged_apis),
                    "added": added,
                    "updated": updated,
                    "api_path": api_path,
                    "deleted": found.get("deleted") is True,
                }
            )

        _save_deps_root(root)
        catalog = DependencyCatalog.load(deps_json)
        append_log("import_deps_excel", {"file": _as_str(file.filename).strip(), "system_count": len(results)})
        return {"ok": True, "system_count": len(results), "systems": results, "deps_file": str(deps_json)}

    def _import_swagger_core(url: str, body: dict, resume_key: str | None = None) -> dict:
        nonlocal catalog
        fetch_headers = _build_fetch_headers(body)
        spec = _fetch_openapi(url, fetch_headers)
        if spec is None:
            return {"ok": False, "error": "failed to fetch or parse swagger/openapi/yapi", "url": url}

        info = spec.get("info") if isinstance(spec.get("info"), dict) else {}
        project = spec.get("project") if isinstance(spec.get("project"), dict) else {}
        title = _as_str(info.get("title")).strip() or _as_str(project.get("name") or project.get("title")).strip() or "External System"
        description = _as_str(info.get("description")).strip() or _as_str(project.get("desc") or project.get("description")).strip() or _as_str(spec.get("desc") or spec.get("description")).strip()
        sys_id = _snake_upper(_as_str(body.get("system_id")).strip() or title)
        if not sys_id:
            return {"ok": False, "error": "failed to derive system_id", "url": url}
        sys_type = _as_str(body.get("type")).strip() or "external_service"
        sys_name = _as_str(body.get("name")).strip() or title

        source = "swagger"
        apis = _extract_openapi_apis(spec)
        if not apis:
            source = "yapi"
            apis = _extract_yapi_apis(spec)
        if not apis:
            return {"ok": False, "error": "no apis found in swagger/openapi/yapi input", "url": url}

        api_file_name = f"{sys_id}.apis.json"
        existing = _read_api_file_obj(api_file_name)
        existing_apis = existing.get("apis") if isinstance(existing.get("apis"), list) else []
        merged_apis, added, updated = _merge_apis(existing_apis, apis)
        api_path = _write_api_file_obj(api_file_name, sys_id, merged_apis)

        root = _load_deps_root()
        found = _upsert_system_root(
            root,
            sys_id,
            sys_name,
            sys_type,
            api_file_name,
            {"description": description, "aliases": [sys_name], "keywords": _keywords_from([sys_name, title])},
        )
        _save_deps_root(root)

        catalog = DependencyCatalog.load(deps_json)
        append_log("import_swagger", {"url": url, "system_id": sys_id, "api_count": len(merged_apis), "source": source})
        if resume_key:
            append_log("import_swagger_batch_item", {"resume_key": resume_key, "url": url, "ok": True, "system_id": sys_id, "source": source, "api_count": len(merged_apis), "added": added, "updated": updated})
        return {
            "ok": True,
            "source": source,
            "system": {"id": sys_id, "name": found.get("name"), "type": found.get("type"), "api_file": api_file_name},
            "api_count": len(merged_apis),
            "added": added,
            "updated": updated,
            "deps_file": str(deps_json),
            "api_file": api_path,
            "url": url,
        }

    @app.post("/api/import_swagger")
    async def import_swagger(request: Request):
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}
        url = _as_str(body.get("url")).strip()
        if not url:
            return JSONResponse({"error": "url is required"}, status_code=400)
        r = _import_swagger_core(url, body)
        if not r.get("ok"):
            return JSONResponse({"error": r.get("error") or "更新失败"}, status_code=400)
        return r

    @app.post("/api/import_swagger_batch")
    async def import_swagger_batch(request: Request):
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}

        urls_in = body.get("urls")
        if isinstance(urls_in, str):
            urls = [x.strip() for x in urls_in.splitlines() if x and x.strip()]
        elif isinstance(urls_in, list):
            urls = [_as_str(x).strip() for x in urls_in if _as_str(x).strip()]
        else:
            urls = []
        if not urls:
            return JSONResponse({"error": "urls is required"}, status_code=400)

        seen: set[str] = set()
        uniq: list[str] = []
        for u in urls:
            if u in seen:
                continue
            seen.add(u)
            uniq.append(u)

        resume_key = _as_str(body.get("resume_key")).strip()
        if not resume_key:
            resume_key = time.strftime("%Y%m%d%H%M%S", time.localtime())

        done = _log_done_urls_for_resume(resume_key)
        results: list[dict] = []
        skipped = 0
        ok = 0
        failed = 0
        for u in uniq:
            if u in done:
                skipped += 1
                results.append({"ok": True, "skipped": True, "url": u})
                continue
            r = _import_swagger_core(u, body, resume_key=resume_key)
            if r.get("ok") is True:
                ok += 1
                results.append(r)
            else:
                failed += 1
                append_log("import_swagger_batch_item", {"resume_key": resume_key, "url": u, "ok": False, "error": r.get("error") or "failed"})
                results.append({"ok": False, "url": u, "error": r.get("error") or "failed"})
        append_log("import_swagger_batch", {"resume_key": resume_key, "total": len(uniq), "ok": ok, "failed": failed, "skipped": skipped})
        return {"ok": True, "resume_key": resume_key, "total": len(uniq), "ok_count": ok, "failed_count": failed, "skipped_count": skipped, "results": results}

    @app.post("/api/front_log")
    async def front_log(request: Request):
        try:
            data = await request.json()
        except Exception:
            data = {}
        if not isinstance(data, dict):
            data = {"value": data}
        append_log("front", data)
        return {"ok": True}

    @app.post("/api/extract")
    async def extract(request: Request, files: list[UploadFile] = File(...)):
        try:
            ip = request.client.host if request.client else ""
        except Exception:
            ip = ""
        scan_id_base = time.strftime("%Y%m%d%H%M%S", time.localtime())
        scan_id = scan_id_base
        append_log("extract_start", {"ip": ip, "scan_id": scan_id, "files": [f.filename for f in (files or []) if f is not None]})

        if cfg.work_dir is not None:
            out_base = (cfg.work_dir / "output").resolve()
        else:
            out_base = (Path.cwd() / "work" / "output").resolve()
        out_dir = (out_base / scan_id).resolve()
        if out_dir.exists():
            i = 1
            while True:
                cand = (out_base / f"{scan_id_base}_{i}").resolve()
                if not cand.exists():
                    scan_id = f"{scan_id_base}_{i}"
                    out_dir = cand
                    break
                i += 1
        out_dir.mkdir(parents=True, exist_ok=True)

        by_file = []
        outputs = []
        for f in files:
            name = f.filename or "upload"
            lower = name.lower()
            if lower.endswith(".doc") or lower.endswith(".docx"):
                suf = ".doc" if lower.endswith(".doc") else ".docx"
                tmp = Path.cwd() / f".spec_dep_upload_{id(f)}{suf}"
                content = await f.read()
                tmp.write_bytes(content)
                try:
                    text = extract_word_text(tmp)
                finally:
                    try:
                        tmp.unlink(missing_ok=True)
                    except Exception:
                        pass
            else:
                raw = await f.read()
                text = raw.decode("utf-8", errors="replace")

            modules = split_text_into_modules(text)
            module_rows = []
            for sec in modules:
                mod = sec.get("module") or "global"
                body = sec.get("text") or ""
                deps = catalog.match_systems(body)
                module_rows.append({"module": mod, "dependencies": deps})
            llm = llm_extract_module_deps(cfg, text, deps_json)
            if llm is not None:
                mods = _dedupe_modules(llm.get("modules") or module_rows)
                by_file.append(
                    {
                        "file": name,
                        "modules": mods,
                        "llm": {"provider": "deepseek", "model": cfg.deepseek_model, "used": True},
                    }
                )
            else:
                by_file.append({"file": name, "modules": _dedupe_modules(module_rows), "llm": {"provider": "deepseek", "model": cfg.deepseek_model, "used": False}})

            base = Path(name).stem or "output"
            per_file = {
                "scan_id": scan_id,
                "requirement_file": name,
                "dependencies_file": str(deps_json),
                "modules": by_file[-1].get("modules") if isinstance(by_file[-1], dict) else _dedupe_modules(module_rows),
                "llm": by_file[-1].get("llm") if isinstance(by_file[-1], dict) else None,
            }
            json_path = (out_dir / f"{base}.module_deps.json").resolve()
            xlsx_path = (out_dir / f"{base}.module_deps.xlsx").resolve()
            json_path.write_text(json.dumps(per_file, ensure_ascii=False, indent=2), encoding="utf-8")
            write_module_dependency_excel(xlsx_path, per_file)
            outputs.append(
                {
                    "file": name,
                    "json": str(json_path),
                    "xlsx": str(xlsx_path),
                    "json_name": json_path.name,
                    "xlsx_name": xlsx_path.name,
                }
            )

        out = {"scan_id": scan_id, "output_dir": str(out_dir), "files": by_file, "dependencies_file": str(deps_json), "outputs": outputs}
        scan_json = (out_dir / "scan.json").resolve()
        try:
            scan_json.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
        append_log("extract_done", {"ip": ip, "scan_id": scan_id, "ok": True, "file_count": len(by_file), "output_dir": str(out_dir)})
        return out

    @app.post("/api/adjust")
    async def adjust(request: Request):
        try:
            body = await request.json()
        except Exception:
            body = {}
        if not isinstance(body, dict):
            body = {}

        scan_id = str(body.get("scan_id") or "").strip()
        instruction = str(body.get("instruction") or "").strip()
        if not scan_id:
            return JSONResponse({"error": "scan_id is required"}, status_code=400)
        if not instruction:
            return JSONResponse({"error": "instruction is required"}, status_code=400)

        out_dir = _resolve_scan_dir(scan_id)
        if out_dir is None:
            if cfg.work_dir is not None:
                out_dir = (cfg.work_dir / "output" / scan_id).resolve()
            else:
                out_dir = (project_root / "work" / "output" / scan_id).resolve()
        scan_json = (out_dir / "scan.json").resolve()
        if not scan_json.is_file():
            return JSONResponse({"error": f"scan.json not found for scan_id={scan_id}"}, status_code=404)

        try:
            scan_obj = json.loads(scan_json.read_text(encoding="utf-8"))
        except Exception:
            scan_obj = {}
        if not isinstance(scan_obj, dict):
            scan_obj = {}

        adjusted = llm_adjust_scan_result(cfg, scan_obj, instruction, deps_json)
        if adjusted is None:
            if cfg.deepseek_api_key is None or not cfg.deepseek_api_key.strip():
                return JSONResponse({"error": "AI 调整不可用：未加载 DEEPSEEK_API_KEY（请检查启动时使用的 config.yaml 或环境变量）"}, status_code=400)
            return JSONResponse({"error": "AI 调整不可用：模型服务请求失败（请检查 deepseek_base_url / 网络 / 模型服务状态）"}, status_code=400)

        files_list = adjusted.get("files")
        if isinstance(files_list, list):
            for it in files_list:
                if not isinstance(it, dict):
                    continue
                it["modules"] = _dedupe_modules(it.get("modules"))

        adjusted["scan_id"] = scan_id
        adjusted["output_dir"] = str(out_dir)
        adjusted_file = (out_dir / "scan.adjusted.json").resolve()
        try:
            adjusted_file.write_text(json.dumps(adjusted, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass

        append_log("adjust_done", {"scan_id": scan_id, "ok": True, "output_dir": str(out_dir)})
        adjusted["adjusted_file"] = str(adjusted_file)
        return adjusted

    @app.get("/api/scan_exists")
    def scan_exists(scan_id: str | None = None):
        sid = (scan_id or "").strip()
        if not sid:
            return {"exists": False}
        found = _resolve_scan_dir(sid)
        if found is not None:
            return {"scan_id": sid, "exists": True, "output_dir": str(found)}
        if cfg.work_dir is not None:
            out_dir = (cfg.work_dir / "output" / sid).resolve()
        else:
            out_dir = (project_root / "work" / "output" / sid).resolve()
        return {"scan_id": sid, "exists": False, "output_dir": str(out_dir)}

    @app.get("/api/scan_result")
    def get_scan_result(scan_id: str):
        sid = (scan_id or "").strip()
        if not sid:
            return JSONResponse({"error": "scan_id is required"}, status_code=400)
        
        out_dir = _resolve_scan_dir(sid)
        if out_dir is None:
            return JSONResponse({"error": f"scan.json not found for scan_id={sid}"}, status_code=404)
        
        scan_json = (out_dir / "scan.json").resolve()
        if not scan_json.is_file():
            return JSONResponse({"error": f"scan.json not found for scan_id={sid}"}, status_code=404)
            
        try:
            scan_obj = json.loads(scan_json.read_text(encoding="utf-8"))
            return scan_obj
        except Exception as e:
            return JSONResponse({"error": f"failed to read scan.json: {e}"}, status_code=500)

    @app.get("/api/output_file")
    def output_file(scan_id: str, name: str):
        sid = (scan_id or "").strip()
        fname = (name or "").strip()
        if not sid or not fname:
            return JSONResponse({"error": "scan_id and name are required"}, status_code=400)
        if "/" in fname or "\\" in fname:
            return JSONResponse({"error": "invalid file name"}, status_code=400)
        lower = fname.lower()
        if not (lower.endswith(".xlsx") or lower.endswith(".json")):
            return JSONResponse({"error": "only .xlsx/.json can be downloaded"}, status_code=400)

        out_dir = _resolve_scan_dir(sid)
        if out_dir is None:
            return JSONResponse({"error": f"scan.json not found for scan_id={sid}"}, status_code=404)
        p = (out_dir / fname).resolve()
        if p.parent != out_dir or not p.is_file():
            return JSONResponse({"error": "file not found"}, status_code=404)
        return FileResponse(str(p))

    @app.get("/")
    def index():
        idx = dist / "index.html"
        if idx.is_file():
            return FileResponse(str(idx))
        return JSONResponse({"error": "frontend not built"}, status_code=404)

    @app.get("/{path:path}")
    def spa_fallback(path: str):
        if path.startswith("api/") or path == "api":
            return JSONResponse({"error": "not found"}, status_code=404)
        candidate = dist / path
        if candidate.is_file():
            return FileResponse(str(candidate))
        idx = dist / "index.html"
        if idx.is_file():
            return FileResponse(str(idx))
        return JSONResponse({"error": "frontend not built"}, status_code=404)

    return app


def _resolve_deps_json(cfg) -> Path:
    if cfg.work_dir is not None:
        dep_dir = (cfg.work_dir / "input" / "dependencies").resolve()
    else:
        dep_dir = (Path.cwd() / "work" / "input" / "dependencies").resolve()
    if dep_dir.is_dir():
        idx = (dep_dir / "index.json").resolve()
        if idx.is_file():
            return idx
        candidates = [p for p in dep_dir.iterdir() if p.is_file() and p.name.lower().endswith(".json")]
        candidates.sort(key=lambda p: p.name.lower())
        if candidates:
            return candidates[0].resolve()
    raise RuntimeError(f"dependencies json not found under: {dep_dir}")
